import openpyxl

wb = openpyxl.load_workbook(r"client/public/DDTME_BMD - May'26.xlsx", data_only=True)
ws = wb.active

# Print first few rows in detail to see header structure
for i in range(1, 6):
    row_data = [ws.cell(row=i, column=j).value for j in range(1, ws.max_column + 1)]
    print(f"Row {i}: {row_data}")

# Print merged cells info
print("\nMerged cells:")
for mc in ws.merged_cells.ranges:
    print(f"  {mc}")
